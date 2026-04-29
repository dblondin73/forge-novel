"""
Build research/skill-taxonomy-reference.xlsx from a curated skill list
sourced from D&D 5e, major LitRPG series, ESO, WoW, and forge-novel cosmology.

Rerun after editing SKILLS, LIFE_SKILLS, or SOURCES_KEY to regenerate the workbook.

Columns: Skill | Preferred Variant | Category | Sources | Forge Mapping | Notes | Faculty | Counterfeit Of

Faculty: analytical / relational / both / "" — per Codex #95 Two-Faculty Model.
Counterfeit Of: the Conduit Gift this skill is a System-knockoff of (Codex #178) — empty unless applicable.

2026-04-23: added LIFE_SKILLS (Domain 10) catalog and two trailing columns.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).parent / "skill-taxonomy-reference.xlsx"

# ──────────────────────────────────────────────────────────────────────────
# Source key — abbreviations used in the Sources column
# ──────────────────────────────────────────────────────────────────────────

SOURCES_KEY = [
    ("D&D",        "D&D 5e — canonical 18-skill list"),
    ("Land",       "The Land (Chaos Seeds) — Aleron Kong"),
    ("PrimalHunt", "The Primal Hunter — Zogarth"),
    ("HWFWM",      "He Who Fights With Monsters — Shirtaloon"),
    ("DCC",        "Dungeon Crawler Carl — Matt Dinniman"),
    ("Defiance",   "Defiance of the Fall — TheFirstDefier"),
    ("ForgeMstr",  "Forge Master (Tower) — Seth Ring"),
    ("Cradle",     "Cradle — Will Wight (Path/Technique framework, not skill list)"),
    ("PoA",        "Path of Ascension — C. Mantis"),
    ("RtM",        "Road to Mastery — Valerios"),
    ("ESO",        "Elder Scrolls Online"),
    ("WoW",        "World of Warcraft (incl. retired weapon skills)"),
    ("IronDruid",  "The Iron Druid Chronicles — Kevin Hearne (companion bond model)"),
    ("Forge",      "Forge Novel — locked Codex entity"),
]

# ──────────────────────────────────────────────────────────────────────────
# Skills — (name, preferred_variant, category, sources, forge_mapping, notes)
#
# The `preferred_variant` column captures the editorial pick of which form
# of the name to use in prose / HUD readouts. Kept alongside `name` so the
# full compound is preserved for disambiguation.
# ──────────────────────────────────────────────────────────────────────────

SKILLS = [
    # ── Combat — Weapons ─────────────────────────────────────────────────
    ("One-Handed",                       "One-Handed",                  "Combat — Weapons", "ESO, WoW, Land",                         "",                               "Generic one-handed melee"),
    ("Two-Handed",                       "Two-Handed",                  "Combat — Weapons", "ESO, WoW, Land",                         "",                               "Greatswords, mauls, greataxes"),
    ("Sword Mastery",                    "Sword Mastery",               "Combat — Weapons", "Land, Defiance",                         "",                               "Classic LitRPG weapon skill"),
    ("Axe Mastery",                      "Axe Mastery",                 "Combat — Weapons", "WoW, Land",                              "",                               "Two-handed or dual-wield"),
    ("Mace Mastery",                     "Mace Mastery",                "Combat — Weapons", "WoW, Land",                              "",                               "Blunt force / armor-crushing"),
    ("Dagger Mastery",                   "Dagger Mastery",              "Combat — Weapons", "WoW, Land",                              "",                               "Rogue / assassin pairing"),
    ("Spear Mastery",                    "Spear Mastery",               "Combat — Weapons", "PrimalHunt, Land",                       "",                               "Jake's signature line — Shape of the Spear evolves"),
    ("Bow / Archery",                    "Archery",                     "Combat — Weapons", "ESO, WoW, Land, D&D",                    "",                               "Ranged physical"),
    ("Crossbow / Marksmanship",          "Marksmanship",                "Combat — Weapons", "WoW, Land",                              "Nate (improvised crossbow)",     "Tutorial loadout weapon (Codex #104)"),
    ("Unarmed / Martial Arts",           "Martial Arts",                "Combat — Weapons", "RtM, WoW",                               "",                               "Jack Rust's fist-only build"),
    ("Shield (Block / Bash)",            "Shield (Block / Bash)",       "Combat — Weapons", "ESO, WoW",                               "",                               "Defensive + offensive"),
    ("Dual Wield",                       "Dual Wield",                  "Combat — Weapons", "ESO, WoW, Land",                         "",                               "Two light weapons"),
    ("Polearm / Halberd",                "Polearm",                     "Combat — Weapons", "WoW",                                    "",                               "Reach weapon"),
    ("Throwing Weapons",                 "Throwing Weapons",            "Combat — Weapons", "Land, WoW",                              "",                               "Thrown axes, knives, javelins"),
    ("Hammer (reinforced / war)",        "Hammer (reinforced / war)",   "Combat — Weapons", "Land, ForgeMstr",                        "Nate (reinforced hammer)",       "Tutorial loadout — also crafting tool (Codex #104)"),

    # ── Combat — Magic Schools / Elemental ──────────────────────────────
    ("Fire / Pyromancy",                 "Pyromancy",                   "Combat — Magic",   "ESO, WoW, Land",                         "",                               "Destruction Staff (Fire) line in ESO"),
    ("Frost / Cryomancy",                "Cryomancy",                   "Combat — Magic",   "ESO, WoW, Defiance",                     "",                               "Cold-element damage + slow"),
    ("Lightning / Storm Calling",        "Lightning",                   "Combat — Magic",   "ESO (Sorcerer), Land, PrimalHunt",       "",                               "Shock damage, fast cast"),
    ("Earth / Stone",                    "Earth",                       "Combat — Magic",   "ESO (Warden), Land",                     "",                               "Defensive + ranged rock"),
    ("Wind / Air",                       "Air",                         "Combat — Magic",   "Land, Defiance",                         "",                               "Mobility + ranged disruption"),
    ("Water / Ice (non-Frost)",          "Water",                       "Combat — Magic",   "Land",                                   "",                               "Tidal / healing adjacency"),
    ("Arcane / Pure Mana",               "Arcane",                      "Combat — Magic",   "WoW, D&D, PrimalHunt",                   "",                               "Raw energy, no elemental affinity"),
    ("Holy / Light / Divine",            "Light",                       "Combat — Magic",   "WoW, D&D",                               "",                               "Damage undead, buff allies — System counterfeit of Conduit"),
    ("Shadow / Dark",                    "Shadow",                      "Combat — Magic",   "ESO (Necromancer), WoW, Land",           "",                               "Life-drain, fear, curses"),
    ("Necromancy",                       "Necromancy",                  "Combat — Magic",   "ESO, WoW, Land",                         "",                               "Raise undead, soul manipulation"),
    ("Nature / Druidic",                 "Druidic",                     "Combat — Magic",   "ESO (Warden), WoW (Druid)",              "",                               "Plant/animal-based magic"),
    ("Summoning",                        "Summoning",                   "Combat — Magic",   "HWFWM, WoW (Warlock)",                   "",                               "Call in minions or allies"),
    ("Restoration / Healing",            "Healing",                     "Combat — Magic",   "ESO, WoW, D&D",                          "",                               "HoT, direct heal, shield"),
    ("Illusion / Enchantment",           "Enchantment",                 "Combat — Magic",   "D&D, WoW (old)",                         "",                               "Mind control, disguise, charm"),

    # ── Combat — Tactical / Support ─────────────────────────────────────
    ("Leadership / Party Awareness",     "Leadership",                  "Combat — Tactical","WoW, D&D-adj",                           "Marcus (Vanguard #99)",          "Sees health/positioning/threat for allies"),
    ("Tactics / Battle Analysis",        "Tactics",                     "Combat — Tactical","Land, PrimalHunt",                       "",                               "Hunter's Insight-style skill"),
    ("Buff / Aura / Banner",             "Buff",                        "Combat — Tactical","WoW (Paladin), ESO (Templar)",           "",                               "Party-wide stat boosts"),
    ("Debuff / Curse / Hex",             "Debuff",                      "Combat — Tactical","WoW (Warlock), ESO (Nightblade)",        "",                               "Weaken enemies before combat"),
    ("Resurrection / Revive",            "Revive",                      "Combat — Tactical","WoW, Land, ESO",                         "",                               "Mid-combat or post-combat revive"),
    ("Threat / Taunt",                   "Taunt",                       "Combat — Tactical","WoW, ESO",                               "",                               "Tank aggro management"),

    # ── Perception & Awareness ──────────────────────────────────────────
    ("Perception",                       "Perception",                  "Perception",       "D&D",                                    "",                               "See, hear, notice"),
    ("Investigation",                    "Investigation",               "Perception",       "D&D",                                    "",                               "Deduce from clues, search scene"),
    ("Insight",                          "Insight",                     "Perception",       "D&D",                                    "",                               "Read motives and intent"),
    ("Danger Sense",                     "Danger Sense",                "Perception",       "PrimalHunt, HWFWM, Land",                "",                               "Precognitive threat warning"),
    ("Enhanced Senses",                  "Enhanced Senses",             "Perception",       "PrimalHunt, IronDruid",                  "Rex / Judge (Companion Resonance #88)","Heightened sight/hearing/smell"),
    ("Tracking",                         "Tracking",                    "Perception",       "D&D, WoW (Hunter), ESO",                 "",                               "Follow trail, identify quarry"),
    ("Analyze / Examine",                "Analyze",                     "Perception",       "PrimalHunt, Land, ForgeMstr",            "",                               "Read hidden properties of objects/creatures"),
    ("Structural Analysis",              "Structural Analysis",         "Perception",       "Forge (glitched)",                       "Nate (glitched Engineer #98)",   "Sees actual architecture — no clean source parallel"),
    ("Mana Sight",                       "Mana Sight",                  "Perception",       "PrimalHunt, Defiance",                   "",                               "See mana flow, enchantments, bonds"),

    # ── Stealth & Subtlety ──────────────────────────────────────────────
    ("Stealth / Sneak",                  "Stealth",                     "Stealth",          "D&D, ESO, WoW",                          "",                               "Move unseen/unheard"),
    ("Sleight of Hand",                  "Sleight of Hand",             "Stealth",          "D&D, ESO (Legerdemain), WoW",            "",                               "Pickpocketing, palming"),
    ("Lockpicking",                      "Lockpicking",                 "Stealth",          "ESO, Land, WoW",                         "",                               "Mechanical locks, often tool-gated"),
    ("Disguise",                         "Disguise",                    "Stealth",          "D&D, Land",                              "",                               "Alter appearance"),
    ("Trap Detection / Disarm",          "Disarm",                      "Stealth",          "D&D, WoW (Rogue), Land",                 "",                               "Rogue-adjacent"),

    # ── Social ──────────────────────────────────────────────────────────
    ("Persuasion",                       "Persuasion",                  "Social",           "D&D, ESO",                               "",                               "Honest negotiation"),
    ("Intimidation",                     "Intimidation",                "Social",           "D&D, ESO",                               "",                               "Hostile persuasion"),
    ("Deception / Bluff",                "Deception",                   "Social",           "D&D, Land",                              "",                               "Lies, cons, fast-talk"),
    ("Performance",                      "Performance",                 "Social",           "D&D, Land (Bard)",                       "",                               "Music, acting, oratory"),
    ("Animal Handling",                  "Animal Handling",             "Social",           "D&D, WoW (Hunter Pet), Land",            "Nate (Rex + Judge)",             "Natural fit — real-world rancher skill"),
    ("Diplomacy",                        "Diplomacy",                   "Social",           "D&D-adj",                                "",                               "Formal faction negotiation"),
    ("Leadership (social, non-combat)",  "Leadership (social)",         "Social",           "WoW (Guild)",                            "",                               "Faction/guild management"),

    # ── Physical & Mobility ─────────────────────────────────────────────
    ("Athletics",                        "Athletics",                   "Physical",         "D&D, Land",                              "",                               "Climb, jump, swim, grapple"),
    ("Acrobatics / Tumbling",            "Acrobatics",                  "Physical",         "D&D, ESO (Medium Armor)",                "",                               "Balance, dodge, roll"),
    ("Sprint / Endurance",               "Endurance",                   "Physical",         "WoW (old), ESO (Stamina)",               "",                               "Sustained movement"),
    ("Swimming / Diving",                "Swimming",                    "Physical",         "WoW, D&D (Athletics)",                   "",                               "Underwater combat / breath-holding"),
    ("Climbing",                         "Climbing",                    "Physical",         "D&D (Athletics)",                        "",                               "Vertical traversal"),
    ("Heavy Armor",                      "Heavy Armor",                 "Physical",         "ESO, WoW (retired)",                     "",                               "Defensive, stamina-heavy"),
    ("Medium Armor",                     "Medium Armor",                "Physical",         "ESO",                                    "",                               "Balanced defense/mobility"),
    ("Light Armor",                      "Light Armor",                 "Physical",         "ESO",                                    "",                               "Caster/rogue — high evasion"),
    ("Resilience / Fortitude",           "Resilience",                  "Physical",         "WoW, PrimalHunt",                        "",                               "Resist disease, poison, status"),

    # ── Gathering / Survival ────────────────────────────────────────────
    ("Survival",                         "Survival",                    "Gathering",        "D&D",                                    "",                               "Wilderness navigation, foraging"),
    ("Nature",                           "Nature",                      "Gathering",        "D&D",                                    "",                               "Plant/animal lore"),
    ("Mining / Ore Extraction",          "Mining",                      "Gathering",        "WoW, ESO, Land",                         "",                               "Metal + gem nodes"),
    ("Scavenging",                       "Scavenging",                  "Gathering",        "Defiance",                               "",                               "Ruin/battlefield salvage"),
    ("Herbalism / Foraging",             "Herbalism",                   "Gathering",        "WoW, Land",                              "",                               "Plants for alchemy/cooking"),
    ("Skinning / Butchering",            "Skinning",                    "Gathering",        "WoW, Land",                              "",                               "Hide/meat/organ harvest"),
    ("Fishing",                          "Fishing",                     "Gathering",        "WoW, ESO (Provisioning), Land",          "",                               "Water-based food"),
    ("Hunting / Trapping",               "Hunting",                     "Gathering",        "PrimalHunt, WoW (Hunter)",               "",                               "Active creature pursuit"),
    ("Archaeology / Excavation",         "Archaeology",                 "Gathering",        "WoW",                                    "",                               "Ruin digs"),
    ("Electronic Intrusion",             "Electronic Intrusion",        "Gathering",        "Forge (adapted)",                        "",                               "Hack consoles — fits Engineer vibe"),
    ("Identify / Lesser Identify",       "Identify",                    "Gathering",        "Land, PrimalHunt, HWFWM, ESO",           "",                               "Reveal item properties"),
    ("Appraise",                         "Appraise",                    "Gathering",        "Land, D&D-adj",                          "Josie (Appraiser #103)",         "Valuation + material/quality read"),

    # ── Crafting / Production ───────────────────────────────────────────
    ("Blacksmithing / Smithing",         "Blacksmithing",               "Crafting",         "WoW, ESO, Land, ForgeMstr",              "",                               "Central to Forge Master"),
    ("Armorsmithing",                    "Armorsmithing",               "Crafting",         "ESO",                                    "",                               "Heavy/medium armor production"),
    ("Weaponsmithing",                   "Weaponsmithing",              "Crafting",         "ESO",                                    "",                               "Weapon-specific splits"),
    ("Woodworking / Carpentry",          "Woodworking",                 "Crafting",         "ESO, Land",                              "",                               "Bows, staves, shields"),
    ("Fletching / Bowyer",               "Fletching",                   "Crafting",         "Land, D&D (tool)",                       "",                               "Arrows + bows"),
    ("Leatherworking / Tanning",         "Leatherworking",              "Crafting",         "WoW, Land",                              "",                               "Hide-to-armor pipeline"),
    ("Tailoring / Clothing",             "Tailoring",                   "Crafting",         "WoW, ESO, Land",                         "",                               "Cloth armor, bags"),
    ("Jewelcrafting / Jewelry",          "Jewelcrafting",               "Crafting",         "WoW, ESO, Land",                         "",                               "Stat-boosting accessories"),
    ("Enchanting / Runesmithing",        "Runesmithing",                "Crafting",         "WoW, ESO, Land",                         "",                               "Magical item empowerment"),
    ("Inscription / Scribing",           "Inscription",                 "Crafting",         "WoW, Land",                              "",                               "Scrolls, glyphs, tomes"),
    ("Alchemy / Potion Brewing",         "Alchemy",                     "Crafting",         "WoW, ESO, Land, D&D",                    "",                               "Potions, elixirs, poisons"),
    ("Artifice",                         "Artifice",                    "Crafting",         "Land",                                   "",                               "Magical/Force-focus crafting"),
    ("Engineering",                      "Engineering",                 "Crafting",         "WoW",                                    "",                               "Gadgets, traps, vehicles"),
    ("Provisioning / Cooking",           "Cooking",                     "Crafting",         "ESO, WoW, Land",                         "",                               "Buff food, morale"),
    ("Field Repair",                     "Field Repair",                "Crafting",         "Forge, WoW (old)",                       "Nate (Engineer #98)",            "Reads materials through touch — canonical skill"),

    # ── LitRPG-Specific Meta Skills ─────────────────────────────────────
    ("Mana Manipulation / Mana Sense",   "Mana Manipulation",           "LitRPG Meta",      "PrimalHunt, Defiance",                   "",                               "Active mana control, not just pool"),
    ("Soul Sense / Spirit Sight",        "Spirit Sight",                "LitRPG Meta",      "HWFWM, Cradle",                          "",                               "Read souls, essence, alignment"),
    ("Class Awakening / Advancement",    "Advancement",                 "LitRPG Meta",      "All LitRPG sources",                     "",                               "Threshold upgrade to subclass"),
    ("Title Acquisition",                "Title Acquisition",           "LitRPG Meta",      "PrimalHunt, Land, HWFWM",                "",                               "Passive bonuses from named achievements"),
    ("Domain / Territory",               "Domain",                      "LitRPG Meta",      "Cradle, HWFWM, PoA",                     "Cornerstone (#149, in-progress)","Area-of-influence mechanic"),
    ("Skill Evolution / Fusion",         "Skill Evolution",             "LitRPG Meta",      "PrimalHunt, HWFWM",                      "",                               "Skills merge or transform at higher ranks"),
    ("Companion Bond / Beast Tame",      "Companion Bond / Beast Tame", "LitRPG Meta",      "WoW, ESO, IronDruid, ForgeMstr",         "Nate (Rex + Judge pack-bond)",   "Iron Druid's Atticus/Oberon is closest reference"),
    ("Patron Bond",                      "Patron Bond",                 "LitRPG Meta",      "Forge (#97), HWFWM (adj)",               "System hook (Codex #97)",        "Forge canon — binds deeper each tier"),
    ("Conduit Channel",                  "Conduit Channel",             "LitRPG Meta",      "Forge (#91)",                            "Nate (true tier)",               "Conduit power is flow, not pool (Codex #93)"),
]

# ──────────────────────────────────────────────────────────────────────────
# Life Skills — Domain 10 (added 2026-04-23)
#
# (name, preferred, category, sources, forge_mapping, notes, faculty, counterfeit_of)
#
# Faculty tags: analytical / relational / both / "" (undecided)
# Counterfeit Of: Conduit Gift this skill's System-version would imitate (empty for most Life Skills)
# ──────────────────────────────────────────────────────────────────────────

LIFE_SKILLS = [
    # ── Home ─────────────────────────────────────────────────────────────
    ("Cooking",               "Cooking",               "Life Skills — Home",       "Forge",  "",                         "Knife Work, Heat Mgmt, Seasoning, Preservation, Butchery. Butchery graduates at C-rank", "both",       ""),
    ("Baking",                "Baking",                "Life Skills — Home",       "Forge",  "",                         "Doughcraft, Yeast Handling, Ovencraft, Pastry, Sourdough Culture",                        "analytical", ""),
    ("Household Management",  "Household Management",  "Life Skills — Home",       "Forge",  "",                         "Cleaning, Laundry, Inventory, Budgeting. Budgeting overlaps Profession",                  "analytical", ""),
    ("Childcare",             "Childcare",             "Life Skills — Home",       "Forge",  "",                         "Infant Care, Development, Discipline, Teaching. Teaching shared Social/Profession",       "both",       ""),
    ("Eldercare",             "Eldercare",             "Life Skills — Home",       "Forge",  "",                         "Mobility Assist, Dignity Care, Medical Advocacy",                                         "relational", ""),
    ("Sewing / Mending",      "Sewing",                "Life Skills — Home",       "Forge",  "",                         "Hand & Machine Stitching, Patching, Alteration. Graduates to Tailoring",                  "analytical", ""),
    ("Gardening",             "Gardening",             "Life Skills — Home",       "Forge",  "",                         "Kitchen Garden, Orchard, Flower, Composting. Graduates to Herbalism",                     "both",       ""),
    ("Home Repair",           "Home Repair",           "Life Skills — Home",       "Forge",  "",                         "Patching, Painting, Fixture Replacement, Sealing. Gateway to Trade",                      "analytical", ""),
    ("Food Preservation",     "Food Preservation",     "Life Skills — Home",       "Forge",  "",                         "Canning, Smoking, Curing, Fermentation, Root Cellar. Graduates from Cooking",             "analytical", ""),
    ("Hospitality",           "Hospitality",           "Life Skills — Home",       "Forge",  "",                         "Hosting, Table Craft, Guest Care, Event Organization",                                    "relational", ""),

    # ── Trade ────────────────────────────────────────────────────────────
    ("Carpentry",             "Carpentry",             "Life Skills — Trade",      "Forge",  "",                         "Framing, Finish, Cabinetry, Joinery. Overlaps Crafting at high rank",                     "analytical", ""),
    ("Masonry",               "Masonry",               "Life Skills — Trade",      "Forge",  "",                         "Stonework, Brick, Mortar, Foundation, Dry-stack",                                         "analytical", ""),
    ("Plumbing",              "Plumbing",              "Life Skills — Trade",      "Forge",  "",                         "Pipe Fitting, Drain Repair, Fixture Install, Well Work. Earth-bleed retains; outer loses", "analytical", ""),
    ("Electrical",            "Electrical",            "Life Skills — Trade",      "Forge",  "",                         "Wiring, Panel, Motor, Signal. Earth-bleed ONLY — outer Labyrinth lacks substrate",        "analytical", ""),
    ("Welding",               "Welding",               "Life Skills — Trade",      "Forge",  "",                         "Stick, MIG, TIG, Torch. Earth-bleed only at modern ranks; blacksmith substitution",       "analytical", ""),
    ("Mechanical Repair",     "Mechanical Repair",     "Life Skills — Trade",      "Forge",  "",                         "Diagnostic, Disassembly, Gasket, Lubrication. Small-Engine graduates",                    "analytical", ""),
    ("Small Engine Repair",   "Small Engine Repair",   "Life Skills — Trade",      "Forge",  "Nate (ranch / farm)",      "Carburetion, Ignition, Compression, Scheduling. Nate-adjacent",                           "analytical", ""),
    ("Farriery",              "Farriery",              "Life Skills — Trade",      "Forge",  "",                         "Trim, Shoe Fit, Hot Shoeing, Hoof Diagnostics. Overlaps Horsemanship",                    "both",       ""),
    ("Leatherworking",        "Leatherworking",        "Life Skills — Trade",      "Forge",  "",                         "Tanning, Cutting, Stitching, Tooling, Dyeing. Crafting graduation",                       "analytical", ""),
    ("Locksmithing",          "Locksmithing",          "Life Skills — Trade",      "Forge",  "",                         "Mechanical Lock, Pin Tumbler, Key Cutting, Forensic. Overlaps Stealth/Perception",        "analytical", ""),

    # ── Profession ───────────────────────────────────────────────────────
    ("Medical Practice",      "Medical Practice",      "Life Skills — Profession", "Forge",  "",                         "First Aid → Nursing → Paramedic → Physician → Surgeon. Diagnosis, Suturing, Pharm, Triage", "both",     ""),
    ("Nursing",               "Nursing",               "Life Skills — Profession", "Forge",  "Ana Torres",               "Patient Care, Medication, Charting, Palliative",                                          "both",       ""),
    ("Veterinary",            "Veterinary",            "Life Skills — Profession", "Forge",  "",                         "Species Knowledge, Surgical, Behavioral, Birth Attendance. Overlaps Horsemanship",        "both",       ""),
    ("Teaching",              "Teaching",              "Life Skills — Profession", "Forge",  "",                         "Curriculum, Assessment, Classroom Mgmt, Mentorship. Overlaps Social/Mind",                "both",       ""),
    ("Counseling",            "Counseling",            "Life Skills — Profession", "Forge",  "",                         "Active Listening, Therapeutic Technique, Crisis Mgmt",                                    "relational", ""),
    ("Bookkeeping",           "Bookkeeping",           "Life Skills — Profession", "Forge",  "",                         "Ledger, Reconciliation, Tax, Audit",                                                       "analytical", ""),
    ("Accounting",            "Accounting",            "Life Skills — Profession", "Forge",  "",                         "Financial Statement, Cost Accounting, Forensic. Bookkeeping graduates here",              "analytical", ""),
    ("Legal Research",        "Legal Research",        "Life Skills — Profession", "Forge",  "",                         "Case Law, Statute, Citation, Argument Construction",                                       "analytical", ""),
    ("Architecture",          "Architecture",          "Life Skills — Profession", "Forge",  "",                         "Drafting, Load Path Design, Materials, Spec Writing",                                     "analytical", ""),
    ("Civil Engineering",     "Civil Engineering",     "Life Skills — Profession", "Forge",  "Nate (native)",            "Grading, Drainage, Load, Materials, Survey Integration",                                  "analytical", ""),
    ("Structural Analysis",   "Structural Analysis",   "Life Skills — Profession", "Forge",  "Nate (Bezalel ceiling)",   "Load Path, Failure Mode, Safety Factor, Field Assessment. Codex #178 amplification target","analytical", ""),
    ("Surveying",             "Surveying",             "Life Skills — Profession", "Forge",  "",                         "Total Station, GPS, Traverse, Boundary. Cartography graduates",                           "analytical", ""),
    ("Cartography",           "Cartography",           "Life Skills — Profession", "Forge",  "",                         "Projection, Symbology, Field Sketch, Topography",                                         "analytical", ""),
    ("Logistics",             "Logistics",             "Life Skills — Profession", "Forge",  "",                         "Scheduling, Routing, Inventory, Supply Chain, Procurement",                               "both",       ""),
    ("Programming",           "Programming",           "Life Skills — Profession", "Forge",  "",                         "Algorithms, System Design, Debugging, Language Mastery. Earth-substrate dependency",      "analytical", ""),

    # ── Arts ─────────────────────────────────────────────────────────────
    ("Musicianship",          "Musicianship",          "Life Skills — Arts",       "Forge",  "",                         "Fingering, Dynamics, Improv, Ensemble. Separate rank per instrument family",              "both",       ""),
    ("Singing",               "Singing",               "Life Skills — Arts",       "Forge",  "",                         "Range, Pitch, Harmony, Projection, Tone. Overlaps Body",                                  "both",       ""),
    ("Composition",           "Composition",           "Life Skills — Arts",       "Forge",  "",                         "Harmony, Melody, Form, Arrangement",                                                       "analytical", ""),
    ("Visual Arts",           "Visual Arts",           "Life Skills — Arts",       "Forge",  "",                         "Drawing, Painting, Sculpture, Composition, Colour",                                       "both",       ""),
    ("Writing",               "Writing",               "Life Skills — Arts",       "Forge",  "",                         "Prose, Poetry, Journalism, Technical",                                                     "both",       ""),
    ("Storytelling",          "Storytelling",          "Life Skills — Arts",       "Forge",  "",                         "Oral Tradition, Pacing, Voice, Memory Palace. Overlaps Social",                           "relational", ""),
    ("Acting / Theatrical",   "Acting",                "Life Skills — Arts",       "Forge",  "",                         "Character, Voice, Stage Presence, Cold Read. Overlaps Social",                             "relational", ""),
    ("Woodcarving",           "Woodcarving",           "Life Skills — Arts",       "Forge",  "",                         "Relief, In-the-round, Whittling, Chip Carving. Overlaps Carpentry",                       "analytical", ""),
    ("Poetry / Bardic",       "Bardic",                "Life Skills — Arts",       "Forge",  "",                         "Meter, Rhyme, Memorization, Performance. Overlaps Social",                                "both",       ""),

    # ── Body ─────────────────────────────────────────────────────────────
    ("Running",               "Running",               "Life Skills — Body",       "Forge",  "",                         "Sprinting, Distance, Trail, Pacing",                                                       "analytical", ""),
    ("Swimming",              "Swimming",              "Life Skills — Body",       "Forge",  "",                         "Strokes, Open Water, Diving, Lifesaving. Lifesaving overlaps Medical",                    "analytical", ""),
    ("Climbing",              "Climbing",              "Life Skills — Body",       "Forge",  "",                         "Bouldering, Top Rope, Trad, Free Solo, Ice",                                              "analytical", ""),
    ("Horsemanship",          "Horsemanship",          "Life Skills — Body",       "Forge",  "Nate (core)",              "Seat, Aids, Ground Work, Jumping, Ranch Work. Relational-faculty overlap (pack-bond)",    "both",       ""),
    ("Driving",               "Driving",               "Life Skills — Body",       "Forge",  "",                         "Auto, Motorcycle, Off-road, Trailer, Tractor. Earth-bleed only at modern ranks",          "analytical", ""),
    ("Fishing",               "Fishing",               "Life Skills — Body",       "Forge",  "",                         "Baitcasting, Fly, Net, Ice, Spear. Gathering graduation path",                            "analytical", ""),
    ("Hunting (woodcraft)",   "Hunting",               "Life Skills — Body",       "Forge",  "",                         "Stalking, Still-hunt, Tracking Sign, Field Dress. Overlaps Perception + Gathering",       "both",       ""),
    ("Orienteering",          "Orienteering",          "Life Skills — Body",       "Forge",  "",                         "Map, Compass, Terrain Association, Pace Count. Overlaps Perception",                      "analytical", ""),
    ("Martial Arts (civil)",  "Martial Arts",          "Life Skills — Body",       "Forge",  "",                         "Forms, Sparring, Breathing, Meditation. Distinct from System-granted",                    "both",       ""),

    # ── Mind ─────────────────────────────────────────────────────────────
    ("Literacy",              "Literacy",              "Life Skills — Mind",       "Forge",  "",                         "Reading, Writing, Transcription. Rank per writing system",                                "analytical", ""),
    ("Numeracy",              "Numeracy",              "Life Skills — Mind",       "Forge",  "",                         "Arithmetic, Algebra, Statistics, Calculus",                                                "analytical", ""),
    ("Memory Techniques",     "Memory Techniques",     "Life Skills — Mind",       "Forge",  "",                         "Method of Loci, Mnemonic, Rote, Chunking",                                                "analytical", ""),
    ("Analysis",              "Analysis",              "Life Skills — Mind",       "Forge",  "Nate (maker's eye)",       "Decomposition, Pattern, Synthesis, Critique",                                             "analytical", ""),
    ("Linguistics",           "Linguistics",           "Life Skills — Mind",       "Forge",  "",                         "Per-language Fluency, Phonology, Etymology, Translation. Rank per language",              "both",       ""),
    ("Research",              "Research",              "Life Skills — Mind",       "Forge",  "",                         "Source Evaluation, Citation, Archives, Interview",                                        "analytical", ""),
    ("Strategy / Planning",   "Strategy",              "Life Skills — Mind",       "Forge",  "",                         "Scenario, Resource Allocation, Contingency",                                              "analytical", ""),
    ("History / Lore",        "Lore",                  "Life Skills — Mind",       "Forge",  "",                         "Era Knowledge, Source Criticism, Chronology",                                             "analytical", ""),
    ("Celestial Navigation",  "Celestial Navigation",  "Life Skills — Mind",       "Forge",  "",                         "Star ID, Sextant, Almanac, Dead Reckoning. Overlaps Perception",                          "analytical", ""),
    ("Comparative Religion",  "Comparative Religion",  "Life Skills — Mind",       "Forge",  "",                         "Tradition, Text, Practice. PROSE: 'studied the traditions' only",                         "both",       ""),
    ("Philosophy / Logic",    "Logic",                 "Life Skills — Mind",       "Forge",  "",                         "Argumentation, Formal Logic, Rhetoric, Ethics",                                           "analytical", ""),
    ("Codebreaking",          "Codebreaking",          "Life Skills — Mind",       "Forge",  "",                         "Cipher, Frequency, Language Attack, Steganography",                                       "analytical", ""),

    # ── Social ───────────────────────────────────────────────────────────
    ("Conversation",          "Conversation",          "Life Skills — Social",     "Forge",  "",                         "Listening, Mirroring, Small Talk, Escalation. Foundation for Negotiation",                "relational", ""),
    ("Negotiation",           "Negotiation",           "Life Skills — Social",     "Forge",  "",                         "Positioning, BATNA, Concession Trading, Anchor",                                          "both",       ""),
    ("Public Speaking",       "Public Speaking",       "Life Skills — Social",     "Forge",  "Josie",                    "Delivery, Audience Read, Memory, Crowd Pacing",                                           "both",       ""),
    ("Persuasion",            "Persuasion",            "Life Skills — Social",     "Forge",  "Josie",                    "Framing, Social Proof, Reciprocity, Ethos",                                               "relational", ""),
    ("Deception",             "Deception",             "Life Skills — Social",     "Forge",  "",                         "Cover, Misdirection, Body Control, Story Consistency. Dual-use",                          "both",       ""),
    ("Leadership",            "Leadership",            "Life Skills — Social",     "Forge",  "Nate (ranch-scale)",       "Direction, Delegation, Morale, Accountability. Civil/social — distinct from Command",     "both",       ""),
    ("Mentorship",            "Mentorship",            "Life Skills — Social",     "Forge",  "",                         "Assessment, Feedback, Challenge Calibration, Patience",                                    "relational", ""),
    ("Networking",            "Networking",            "Life Skills — Social",     "Forge",  "",                         "Introduction, Follow-up, Referral, Relationship Upkeep",                                  "relational", ""),
    ("Body Language Reading", "Body Language Reading", "Life Skills — Social",     "Forge",  "",                         "Micro-expression, Posture, Congruence, Baseline. Overlaps Perception",                    "relational", ""),
    ("Mediation",             "Mediation",             "Life Skills — Social",     "Forge",  "",                         "Neutrality, Interest Surfacing, Face-saving, Option Framing",                             "relational", ""),
    ("Auctioneering",         "Auctioneering",         "Life Skills — Social",     "Forge",  "Josie (core)",             "Cadence, Crowd Work, Bid Reading, Chant",                                                 "both",       ""),
]

# ──────────────────────────────────────────────────────────────────────────
# Workbook build
# ──────────────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CATEGORY_COLORS = {
    # Domains 1–9 (existing)
    "Combat — Weapons":        "FCE4D6",
    "Combat — Magic":          "D6E5FC",
    "Combat — Tactical":       "E2EFDA",
    "Perception":              "FFF2CC",
    "Stealth":                 "E4E0EC",
    "Social":                  "FFE5D9",
    "Physical":                "DEEBF7",
    "Gathering":               "E7E6E6",
    "Crafting":                "F8CBAD",
    "LitRPG Meta":             "D9D2E9",
    # Domain 10 — Life Skills (2026-04-23)
    "Life Skills — Home":       "FFF8DC",
    "Life Skills — Trade":      "F5DEB3",
    "Life Skills — Profession": "E0F2D9",
    "Life Skills — Arts":       "F4D9F0",
    "Life Skills — Body":       "D6EAF8",
    "Life Skills — Mind":       "E8D5F2",
    "Life Skills — Social":     "FFE4D1",
}

# Normalize both lists into 8-tuples:
# (name, preferred, category, sources, forge_mapping, notes, faculty, counterfeit_of)
ALL_SKILLS = [s + ("", "") for s in SKILLS] + LIFE_SKILLS
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

wb = Workbook()

# ── Master sheet ────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Master"
headers = ["Skill", "Preferred Variant", "Category", "Sources", "Forge Mapping", "Notes", "Faculty", "Counterfeit Of"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = THIN_BORDER

for i, (name, preferred, cat, sources, mapping, notes, faculty, counterfeit) in enumerate(ALL_SKILLS, start=2):
    ws.cell(row=i, column=1, value=name)
    ws.cell(row=i, column=2, value=preferred)
    ws.cell(row=i, column=3, value=cat)
    ws.cell(row=i, column=4, value=sources)
    ws.cell(row=i, column=5, value=mapping)
    ws.cell(row=i, column=6, value=notes)
    ws.cell(row=i, column=7, value=faculty)
    ws.cell(row=i, column=8, value=counterfeit)
    row_fill = PatternFill(
        start_color=CATEGORY_COLORS.get(cat, "FFFFFF"),
        end_color=CATEGORY_COLORS.get(cat, "FFFFFF"),
        fill_type="solid",
    )
    for col in range(1, 9):
        c = ws.cell(row=i, column=col)
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.fill = row_fill

widths = {1: 32, 2: 28, 3: 26, 4: 36, 5: 30, 6: 52, 7: 14, 8: 22}
for col, w in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:H{len(ALL_SKILLS) + 1}"

# ── Sources Key sheet ──────────────────────────────────────────────────
key = wb.create_sheet("Sources Key")
key_headers = ["Abbrev", "Full Name"]
for col, h in enumerate(key_headers, 1):
    cell = key.cell(row=1, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = THIN_BORDER

for i, (abbrev, full) in enumerate(SOURCES_KEY, start=2):
    key.cell(row=i, column=1, value=abbrev).border = THIN_BORDER
    key.cell(row=i, column=2, value=full).border = THIN_BORDER

key.column_dimensions["A"].width = 14
key.column_dimensions["B"].width = 72
key.freeze_panes = "A2"

# ── Category sheets (quick-scan: Preferred | Faculty | Forge Mapping | Notes) ────
for cat in sorted({s[2] for s in ALL_SKILLS}):
    sheet_name = cat.replace("—", "-")[:31]
    cs = wb.create_sheet(sheet_name)
    cs.cell(row=1, column=1, value="Preferred Variant").font = HEADER_FONT
    cs.cell(row=1, column=2, value="Faculty").font = HEADER_FONT
    cs.cell(row=1, column=3, value="Forge Mapping").font = HEADER_FONT
    cs.cell(row=1, column=4, value="Notes").font = HEADER_FONT
    for col in range(1, 5):
        c = cs.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left")
        c.border = THIN_BORDER

    row = 2
    fill = PatternFill(
        start_color=CATEGORY_COLORS.get(cat, "FFFFFF"),
        end_color=CATEGORY_COLORS.get(cat, "FFFFFF"),
        fill_type="solid",
    )
    for name, preferred, c_cat, _, mapping, notes, faculty, _counterfeit in ALL_SKILLS:
        if c_cat != cat:
            continue
        cs.cell(row=row, column=1, value=preferred)
        cs.cell(row=row, column=2, value=faculty)
        cs.cell(row=row, column=3, value=mapping)
        cs.cell(row=row, column=4, value=notes)
        for col in range(1, 5):
            cc = cs.cell(row=row, column=col)
            cc.border = THIN_BORDER
            cc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cc.fill = fill
        row += 1

    cs.column_dimensions["A"].width = 32
    cs.column_dimensions["B"].width = 14
    cs.column_dimensions["C"].width = 32
    cs.column_dimensions["D"].width = 58
    cs.freeze_panes = "A2"

wb.save(OUT)
print(
    f"Wrote {OUT} — {len(ALL_SKILLS)} skills across {len({s[2] for s in ALL_SKILLS})} categories "
    f"({len(SKILLS)} domains 1-9 + {len(LIFE_SKILLS)} Life Skills)"
)
